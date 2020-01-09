from pprint import pprint
from datetime import datetime
import openpyxl
from openpyxl.styles import Alignment
import xlrd
import pyodbc


conn = pyodbc.connect(r'Driver={Microsoft Access Driver (*.mdb, *.accdb)};DBQ=D:\python\Svod\Report.accdb;')
cursor = conn.cursor()
post = {}
book = xlrd.open_workbook('Opl.xls')
sheet = book.sheet_by_index(0)
for i in range(2, sheet.nrows):
    supplier = sheet.cell_value(i, 1)
    if supplier not in post:
        post[supplier] = {'pact': {'number': int(sheet.cell_value(i, 2)),
                                   'date': xlrd.xldate_as_tuple(sheet.cell_value(i, 3), 0)
                                   if isinstance(sheet.cell_value(i, 3), float) else datetime.strptime(
                                       sheet.cell_value(i, 3), '%d.%m.%Y')},
                          'paper': [{'summa': sheet.cell_value(i, 4),
                                     'date': xlrd.xldate_as_tuple(sheet.cell_value(i, 5), 0)
                                     if isinstance(sheet.cell_value(i, 5), float) else datetime.strptime(
                                         sheet.cell_value(i, 5), '%d.%m.%Y'),
                                     'number': int(sheet.cell_value(i, 6))}]}
    else:
        post[supplier]['paper'].append({'summa': sheet.cell_value(i, 4),
                                        'date': xlrd.xldate_as_tuple(sheet.cell_value(i, 5), 0)
                                        if isinstance(sheet.cell_value(i, 5), float) else datetime.strptime(
                                            sheet.cell_value(i, 3), '%d.%m.%Y'),
                                        'number': int(sheet.cell_value(i, 6))})
    s = sheet.cell_value(i, 5)
    print(type(s))
    if isinstance(s, float):
        cursor.execute(f'''
    
        insert into opl (Nom_pact, Sum, Date_pact, Nomer)
        values({sheet.cell_value(i, 2)}, {sheet.cell_value(i, 4)}, {s}, {sheet.cell_value(i, 6)})
    
        ''')
    else:
        cursor.execute(f'''

                insert into opl (Nom_pact, Sum, Date_pact, Nomer)
                values({sheet.cell_value(i, 2)}, {sheet.cell_value(i, 4)}, '{s}', {sheet.cell_value(i, 6)})

                ''')
    conn.commit()

book = xlrd.open_workbook('Pos.xlsx')
sheet = book.sheet_by_index(0)

for i in range(1, sheet.nrows):
    data = sheet.cell_value(i, 6)
    if isinstance(data, str):
        a = ''
        for char in data:
            if char.isdigit() or char == '.':
                a += char
        data = a
    if sheet.cell_value(i, 4) in post:
        post[sheet.cell_value(i, 4)]['pact']['sum'] = float(data)
    else:
        post[sheet.cell_value(i, 4)] = {'pact': {'number': int(sheet.cell_value(i, 2)),
                                                 'date': xlrd.xldate_as_datetime(sheet.cell_value(i, 3), 0)
                                                 if isinstance(sheet.cell_value(i, 3), float)
                                                 else datetime.strftime(sheet.cell_value(i, 3), '%d.%m.%Y'),
                                                 'sum': float(data)},
                                        'paper': [{'summa': 0,
                                                   'date': datetime.now(),
                                                   'number': 0}]}
    s = sheet.cell_value(i, 3)
    print(f'{int(sheet.cell_value(i, 2))}, {s}, {sheet.cell_value(i, 4)}, {float(data)}')
    if isinstance(s, float):
        cursor.execute(f'''
    
                        insert into Pos (Nom_pact, Date_pact, Supplier, Sum)
                        values({int(sheet.cell_value(i, 2))}, {s}, '{sheet.cell_value(i, 4)}', {float(data)})
    
                        ''')
    else:
        cursor.execute(f'''

                            insert into Pos (Nom_pact, Date_pact, Supplier, Sum)
                            values({int(sheet.cell_value(i, 2))}, '{s}', '{sheet.cell_value(i, 4)}', {float(data)})

                            ''')
    conn.commit()
sum_month = {}
for key in post:
    sum_month[key] = {}
    for dog in post[key]['paper']:
        if isinstance(dog['date'], tuple):
            month = dog['date'][1]
        else:
            month = dog['date'].month
        if month in sum_month[key]:
            sum_month[key][month] += dog['summa']
        else:
            sum_month[key][month] = dog['summa']

work = openpyxl.Workbook()
sheet = work.active
text = ['План', '№ договра', 'Дата', 'Поставщик', 'Сумма договора без НДС/с НДС(13%)']
for i in range(5):
    sheet.merge_cells(start_row=1, start_column=i+1, end_column=i+1, end_row=2)
    sheet.cell(1, i+1).value = text[i]
    sheet.cell(1, i+1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
sheet.merge_cells('F1:Q1')
sheet['F1'].value = 'Оплаты/списание'
sheet['F1'].alignment = Alignment(horizontal='center')
text = ['январь', 'февраль', 'март', 'апрель', 'май', 'июнь', 'июль',
        'август', 'сентябрь', 'октябрь', 'ноябрь', 'декабрь']
for i in range(6, 18):
    sheet.cell(2, i).value = text[i-6]
    sheet.cell(2, i).alignment = Alignment(horizontal='center')
i = 3
nds = 0
pprint(post)
for key in post:
    nds = round(post[key]['pact']['sum'] + post[key]['pact']['sum']*.13, 2)
    sheet.cell(i, 2).value = post[key]['pact']['number']
    if isinstance(post[key]['pact']['date'], tuple):
        s = post[key]['pact']['date']
        sheet.cell(i, 3).value = f'{s[2]}.{s[1]}.{s[0]}'
    else:
        sheet.cell(i, 3).value = f'{s[2]}.{s[1]}.{s[0]}'
    sheet.cell(i, 4).value = key
    sheet.cell(i, 5).value = f'{post[key]["pact"]["sum"]}/{nds}'
    for mon in sum_month[key]:
        sheet.cell(i, mon+5).value = sum_month[key][mon]
    i += 1

work.save('Pos1.xlsx')


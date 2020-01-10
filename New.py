
import openpyxl
from openpyxl.styles import Alignment
import xlrd
import pyodbc
from os import getcwd


def create_report(opl, pos):
    cur_path = getcwd()
    conn = pyodbc.connect(r'Driver={Microsoft Access Driver (*.mdb, *.accdb)};DBQ=%s\Report.accdb;'%(cur_path))
    cursor = conn.cursor()
    book = xlrd.open_workbook(opl)
    sheet = book.sheet_by_index(0)
    for i in range(2, sheet.nrows):
        s = sheet.cell_value(i, 5)
        if isinstance(s, float):
            cursor.execute(f'''

               insert into opl (Nom_pact, Sum, Date_pact, Nomer)
               values({sheet.cell_value(i, 2)}, {sheet.cell_value(i, 4)}, {s}, {sheet.cell_value(i, 6)})

               ''')
        else:
            cursor.execute(f'''

                       insert into opl (Nom_pact, Sum, Date_pact, Nomer)
                       values({sheet.cell_value(i, 2)}, {sheet.cell_value(i, 4)}, '{s}', {sheet.cell_value(i, 6)})

                       ''')
        conn.commit()
    book = xlrd.open_workbook(pos)
    sheet = book.sheet_by_index(0)
    for i in range(1, sheet.nrows):
        data = sheet.cell_value(i, 6)
        if isinstance(data, str):
            a = ''
            for char in data:
                if char.isdigit() or char == '.':
                    a += char
            data = a
        s = sheet.cell_value(i, 3)
        cursor.execute(f'select 1 from pos where Nom_pact={int(sheet.cell_value(i, 2))}')
        yes = len(cursor.fetchall())
        if not yes:
            if isinstance(s, float):
                cursor.execute(f'''
    
                                   insert into Pos (Nom_pact, Date_pact, Supplier, Sum)
                                   values({int(sheet.cell_value(i, 2))}, {s}, '{sheet.cell_value(i, 4)}', {float(data)})
    
                                   ''')
            else:
                cursor.execute(f'''
    
                                       insert into Pos (Nom_pact, Date_pact, Supplier, Sum)
                                       values({int(sheet.cell_value(i, 2))}, '{s}', '{sheet.cell_value(i, 4)}',
                                                {float(data)})
    
                                       ''')
            conn.commit()

    work = openpyxl.Workbook()
    sheet = work.active
    sheet.title = 'Отчет'
    text = ['План', '№ договра', 'Дата', 'Поставщик', 'Сумма договора без НДС/с НДС(13%)']
    for i in range(5):
        sheet.merge_cells(start_row=1, start_column=i + 1, end_column=i + 1, end_row=2)
        sheet.cell(1, i + 1).value = text[i]
        sheet.cell(1, i + 1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    sheet.merge_cells('F1:Q1')
    sheet['F1'].value = 'Оплаты/списание'
    sheet['F1'].alignment = Alignment(horizontal='center')
    text = ['январь', 'февраль', 'март', 'апрель', 'май', 'июнь', 'июль',
            'август', 'сентябрь', 'октябрь', 'ноябрь', 'декабрь']
    for i in range(6, 18):
        sheet.cell(2, i).value = text[i - 6]
        sheet.cell(2, i).alignment = Alignment(horizontal='center')
    i = 3

    cursor.execute('''
    select o.Nom_pact, o.Sum, o.Date_pact, o.Nomer, p.Date_pact, p.Supplier, p.Sum
    from opl o
    left join pos p on o.Nom_pact = p.Nom_pact
    ''')
    for row in cursor.fetchall():
        nom_pact, sum_paper, date_paper, nom_paper, date_pact, supplier, sum_pact = row
        sheet.cell(i, 2).value = nom_pact
        sheet.cell(i, 3).value = f'{date_paper.day}.{date_paper.month}.{date_paper.year}'
        sheet.cell(i, 4).value = supplier
        sheet.cell(i, 5).value = f'{round(float(sum_pact), 2)} / {round(float(sum_pact) + float(sum_pact)*.13, 2)}'
        sheet.cell(i, date_paper.month + 5).value = sheet.cell(i, date_paper.month + 5).value + sum_paper\
            if sheet.cell(i, date_paper.month + 5).value else sum_paper
        sheet.cell(i, date_paper.month + 5).number_format = '0.00 ₽'
        i += 1

    work.save('Report.xlsx')

create_report('Opl.xls', "Pos.xlsx")

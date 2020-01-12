import tkinter as tk
from tkinter import filedialog as fd
from tkinter import messagebox
import openpyxl
from openpyxl.styles import Alignment
import xlrd
import pyodbc
from os.path import split
from os import getcwd


def openfile():
    opl_text.delete(0, tk.END)
    file_name = fd.askopenfilename()
    opl_text.insert(0, file_name)
    pos_text.focus_set()


def openfile1():
    pos_text.delete(0, tk.END)
    file_name = fd.askopenfilename()
    pos_text.insert(0, file_name)
    b3.focus_set()


def report():
    opl_file = opl_text.get()
    pos_file = pos_text.get()
    if opl_file == '':
        messagebox.showinfo(title='Сводная', message='Необходимо указать\n путь до файла')
        opl_text.focus_set()
    elif pos_file == '':
        messagebox.showinfo(title='Сводная', message='Необходимо указать\n путь до файла')
        pos_text.focus_set()
    elif not (split(pos_file)[1].endswith('xlsx') or split(pos_file)[1].endswith('xls')):
        messagebox.showinfo(title='Сводная', message='Указан файл\n с не тем расширением1')
        pos_text.focus_set()
    elif not (split(opl_file)[1].endswith('xlsx') or split(opl_file)[1].endswith('xls')):
        messagebox.showinfo(title='Сводная', message='Указан файл\n с не тем расширением2')
        opl_text.focus_set()
    else:
        create_report(opl=opl_file, pos=pos_file)


def create_report(opl, pos):
    cur_path = getcwd()
    conn = pyodbc.connect(r'Driver={Microsoft Access Driver (*.mdb, *.accdb)};DBQ=%s\Report.accdb;'%(cur_path))
    cursor = conn.cursor()
    book = xlrd.open_workbook(opl)
    sheet = book.sheet_by_index(0)
    row = sheet.row_values(0)
    opl_row = ['№\nп/п', 'Поставщик', 'Договор №', 'Договор дата', 'Документ', '', '']
    if row == opl_row:
        for i in range(2, sheet.nrows):
            s = sheet.cell_value(i, 5)
            if isinstance(s, float):
                cursor.execute(f'''
    
                   insert into opl (Nom_pact, Sum, Date_pact, Nomer)
                   values({sheet.cell_value(i, 2)}, {sheet.cell_value(i, 4)}, {s}, {sheet.cell_value(i, 6)})
    
                   ''')
            else:
                cursor.execute(f'''
    
                           insert into opl (Nom_pact, Sum, Date_pact, Nomer)
                           values({sheet.cell_value(i, 2)}, {sheet.cell_value(i, 4)}, '{s}', {sheet.cell_value(i, 6)})
    
                           ''')
            conn.commit()
        book = xlrd.open_workbook(pos)
        sheet = book.sheet_by_index(0)
        row = sheet.row_values(0)
        pos_row = ['Наименование товаров, работ, услуг',
                   'Инициатор',
                   'Номер договора',
                   'Дата',
                   'Поставщик',
                   'Статус договора',
                   'Сумма договора']
        if row == pos_row:
            for i in range(1, sheet.nrows):
                data = sheet.cell_value(i, 6)
                if isinstance(data, str):
                    a = ''
                    for char in data:
                        if char.isdigit() or char == '.':
                            a += char
                    data = a
                s = sheet.cell_value(i, 3)
                cursor.execute(f'select 1 from pos where Nom_pact={int(sheet.cell_value(i, 2))}')
                yes = len(cursor.fetchall())
                if not yes:
                    if isinstance(s, float):
                        cursor.execute(f'''
            
                                           insert into Pos (Nom_pact, Date_pact, Supplier, Sum)
                                           values({int(sheet.cell_value(i, 2))}, {s}, '{sheet.cell_value(i, 4)}',
                                                {float(data)})
            
                                           ''')
                    else:
                        cursor.execute(f'''
            
                                               insert into Pos (Nom_pact, Date_pact, Supplier, Sum)
                                               values({int(sheet.cell_value(i, 2))}, '{s}', '{sheet.cell_value(i, 4)}',
                                                        {float(data)})
            
                                               ''')
                    conn.commit()
        else:
            messagebox.showinfo(title='Сводная', message='Неправильная структура\n файла pos.xls')

        work = openpyxl.Workbook()
        sheet = work.active
        sheet.title = 'Отчет'
        text = ['План', '№ договора', 'Дата', 'Поставщик', 'Сумма договора без НДС/с НДС(13%)']
        for i in range(5):
            sheet.merge_cells(start_row=1, start_column=i + 1, end_column=i + 1, end_row=2)
            sheet.cell(1, i + 1).value = text[i]
            sheet.cell(1, i + 1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        sheet.merge_cells('F1:Q1')
        sheet['F1'].value = 'Оплаты/списание'
        sheet['F1'].alignment = Alignment(horizontal='center')
        text = ['январь', 'февраль', 'март', 'апрель', 'май', 'июнь', 'июль',
                'август', 'сентябрь', 'октябрь', 'ноябрь', 'декабрь']
        for i in range(6, 18):
            sheet.cell(2, i).value = text[i - 6]
            sheet.cell(2, i).alignment = Alignment(horizontal='center')
        i = 3
        cursor.execute('select Nom_pact from pos')
        numbers = cursor.fetchall()
        for now in numbers:
            cursor.execute(f'''
            select o.Nom_pact, p.Date_pact, p.Supplier, p.Sum
            from opl o
            left join pos p on o.Nom_pact = p.Nom_pact
            where p.Nom_pact={now[0]}
            ''')
            rows = cursor.fetchall()
            if rows:
                nom_pact, date_pact, supplier, sum_pact = rows[0]
                sheet.cell(i, 2).value = nom_pact
                sheet.cell(i, 3).value = f'{date_pact.day}.{date_pact.month}.{date_pact.year}'
                sheet.cell(i, 4).value = supplier
                sheet.cell(i, 5).value = f'{round(float(sum_pact), 2)} / {round(float(sum_pact) + float(sum_pact)*.13, 2)}'
                cursor.execute(f'select Date_pact, Sum from opl where Nom_pact={nom_pact}')
                opls = cursor.fetchall()
                for opl_item in opls:
                    date_paper, sum_paper = opl_item
                    sheet.cell(i, date_paper.month + 5).value = sheet.cell(i, date_paper.month + 5).value + sum_paper\
                        if sheet.cell(i, date_paper.month + 5).value else sum_paper
                    sheet.cell(i, date_paper.month + 5).number_format = '0.00 ₽'
            else:
                cursor.execute(f'select Nom_pact, Date_pact, Supplier, Sum from pos where Nom_pact={now[0]}')
                post = cursor.fetchall()
                nom_pact, date_pact, supplier, sum_pact = post[0]
                sheet.cell(i, 2).value = nom_pact
                sheet.cell(i, 3).value = f'{date_pact.day}.{date_pact.month}.{date_pact.year}'
                sheet.cell(i, 4).value = supplier
                sheet.cell(i, 5).value = f'{round(float(sum_pact), 2)} / {round(float(sum_pact) + float(sum_pact) * .13, 2)}'
            i += 1

        work.save('Report.xlsx')
        messagebox.showinfo(title='Сводная', message='Отчет готов')
    else:
        messagebox.showinfo(title='Сводная', message='Неправильная структура файла\n opl.xls')


def exit_app():
    root.destroy()


root = tk.Tk()
root.title('Сводная таблица')
root.geometry('600x200')
opl_label = tk.Label(text='Введите расположение файла OPL.xlsx')
opl_label.grid(row=0, column=0)
opl_text = tk.Entry()
opl_text.place(x=5, y=25, width=430)
b1 = tk.Button(text='Открыть', command=openfile)
b1.place(x=450, y=20, width=100)
pos_label = tk.Label(text='Введите расположение файла Pos.xlsx')
pos_label.place(x=5, y=55)
pos_text = tk.Entry()
pos_text.place(x=5, y=75, width=430)
b2 = tk.Button(text='Открыть', command=openfile1)
b2.place(x=450, y=75, width=100)
b3 = tk.Button(text='Отчет', command=report)
b3.place(x=250, y=150, width=100)
b4 = tk.Button(text='Выход', command=exit_app)
b4.place(x=450, y=150, width=100)

root.mainloop()
